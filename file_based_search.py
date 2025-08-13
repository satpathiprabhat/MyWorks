#!/usr/bin/env python3
import os
import pandas as pd
from openpyxl import load_workbook

# ==============================
# CONFIGURATION
# ==============================
FLAT_FILE_PATH = "customers.txt"  # Flat file with values to search
EXCEL_FILES = {
    "file1.xlsx": ["CustomerID", "ColumnA", "ColumnB"],  # file_path : list of column names to check
    "file2.xlsx": ["CustomerID", "ColumnA", "ColumnB"],
    "file3.xlsx": ["CustomerID", "ColumnA", "ColumnB"],
    "file4.xlsx": ["CustomerID", "ColumnA", "ColumnB"],
    "file5.xlsx": ["CustomerID", "ColumnA", "ColumnB"],
}
OUTPUT_FILE = "search_results_excel.csv"


# ==============================
# FUNCTIONS
# ==============================
def load_flat_file(file_path):
    """Load flat file rows into a list of tuples for searching."""
    search_values = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            parts = line.strip().split(",")  # assuming CSV-style flat file
            if len(parts) >= 3:
                search_values.append(tuple(part.strip() for part in parts[:3]))
    print(f"[INFO] Loaded {len(search_values)} rows from flat file.")
    return search_values


def search_in_excel(search_row, file_path, columns_to_check):
    """Search for the row values in the Excel file using streaming mode."""
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.active

    # Map header names to column indexes
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    col_indexes = []
    for col_name in columns_to_check:
        if col_name in header:
            col_indexes.append(header.index(col_name))
        else:
            raise ValueError(f"Column '{col_name}' not found in {file_path}")

    # Iterate rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = tuple(str(row[i]).strip() if row[i] is not None else "" for i in col_indexes)
        if values == search_row:
            return True
    return False


def check_rows_in_excel(search_values, excel_files):
    """Check each search row in all Excel files."""
    results = {}
    for row_values in search_values:
        results[row_values] = {}
        for fname, cols in excel_files.items():
            found = search_in_excel(row_values, fname, cols)
            results[row_values][fname] = found
        print(f"[INFO] Checked row: {row_values}")
    return results


def save_results(results, output_file):
    """Save results to CSV."""
    df = pd.DataFrame.from_dict(results, orient='index')
    df.index = ["|".join(index) for index in df.index]  # Combine search values as key
    df.index.name = "SearchRow"
    df.to_csv(output_file)
    print(f"[INFO] Results saved to {output_file}")


# ==============================
# MAIN SCRIPT
# ==============================
if __name__ == "__main__":
    if not os.path.exists(FLAT_FILE_PATH):
        raise FileNotFoundError(f"Flat file not found: {FLAT_FILE_PATH}")
    for file in EXCEL_FILES:
        if not os.path.exists(file):
            raise FileNotFoundError(f"Excel file not found: {file}")

    # 1. Load flat file search rows
    search_rows = load_flat_file(FLAT_FILE_PATH)

    # 2. Check rows in Excel files
    results = check_rows_in_excel(search_rows, EXCEL_FILES)

    # 3. Save results
    save_results(results, OUTPUT_FILE)

    print("[INFO] Excel search processing complete.")